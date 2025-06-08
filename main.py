#
# 功能模块和实现手法整理
#
# 本脚本旨在通过自动化浏览器（Playwright）抓取X.com（原Twitter）上的推文信息，
# 包括推文发布时间、发布者、推文内容、图片链接以及图片在本地的存储路径。
# 同时，它还能识别并保存唯一的发布者信息及其主页链接。
# 抓取到的数据最终会被整理并导出到Excel文件中，并自动打开日志和结果文件。
#
# --- 主要功能模块 ---
#
# 1. 配置管理 (Configuration):
#    - PROXY: 定义HTTP代理地址，用于Playwright浏览器和httpx库的网络请求，规避地理限制或提高访问稳定性。
#    - LOG_DIR, IMAGE_DIR_BASE, RESULTS_DIR: 定义日志文件、图片存储和Excel结果文件的输出目录。
#    - URL_TARGET_FILE: 新增的配置文件，用于存储待抓取推文的URL列表，实现多目标抓取。
#    - 目录创建: 脚本启动时自动创建所需的日志、图片和结果目录，确保文件能正确保存。
#    - 文件命名: 使用时间戳为日志和Excel文件生成唯一名称，避免文件覆盖，便于历史记录管理。
#
# 2. 日志系统 (Logging Setup):
#    - 使用Python内置的 `logging` 模块，配置日志记录器 `twitter_scraper`。
#    - 日志级别设置为 `INFO`，记录重要操作和信息。
#    - 同时配置文件处理器 (`FileHandler`) 和控制台处理器 (`StreamHandler`)，实现日志同时输出到文件和控制台。
#    - 日志格式化: 定义统一的日志输出格式，包含时间、日志器名称、级别和消息，提高日志可读性。
#    - 目的: 方便跟踪脚本运行状态、调试问题和记录抓取过程中的事件。
#
# 3. URL读取 (URL Reading Helper Function - `read_urls_from_file`):
#    - 目的: 从 `urlTarget.txt` 文件中读取待抓取的目标URL列表。
#    - 实现手法:
#      - 按行读取文件内容，每行视为一个潜在URL。
#      - 对读取到的URL进行基本校验（非空且以"http"开头），确保有效性。
#      - 错误处理: 如果文件不存在或内容为空，会记录相应的错误或警告信息。
#    - 优点: 提高脚本的灵活性和可配置性，无需硬编码目标URL。
#
# 4. 浏览器自动化与数据抓取 (Browser Automation & Data Scraping - `get_illustration`):
#    - 核心模块，负责实际的网页交互和数据提取。
#    - 异步操作: 使用 `asyncio` 和 `playwright.async_api` 实现异步并发抓取，提高效率。
#    - 浏览器上下文管理: 为每个抓取任务创建独立的浏览器上下文，加载 `cookies.json` 中的会话信息，保持登录状态。
#      - Cookie处理: 对 `sameSite` 属性进行兼容性处理，以适应Playwright的要求。
#    - 页面导航: 使用 `page.goto()` 导航到目标URL，设置超时以应对网络问题。
#    - 动态内容加载:
#      - 滚动加载: 循环执行 `window.scrollTo(0, document.body.scrollHeight)` 模拟用户滚动，触发页面加载更多推文。
#      - 元素等待: 使用 `expect(page.locator('article').nth(0)).to_be_visible()` 等待新内容（推文文章）出现，确保页面加载完成。
#    - 数据提取 (使用BeautifulSoup辅助):
#      - 遍历 `article` 元素: 识别并处理页面上的每个推文。
#      - 广告过滤: 根据特定HTML结构（如 `Ad` 文本）跳过广告推文。
#      - 发布时间: 从 `<time>` 元素的 `datetime` 属性中提取，并格式化。
#      - 推文URL: 从 `<time>` 元素的父级链接中提取完整推文地址。
#      - 推文内容: 从 `data-testid="tweetText"` 的 `div` 中提取文本内容。
#      - 图片链接: 从 `data-testid="tweetPhoto"` 中提取 `img` 标签的 `src` 属性，并过滤掉视频缩略图。
#      - 发布者信息: 从 `data-testid="User-Name"` 中提取发布者名称和@Handle，构建发布者主页链接。
#    - 数据去重: 使用 `processed_tweet_urls` 集合来避免重复处理同一条推文，提高效率。
#    - 图片下载:
#      - 使用 `httpx.AsyncClient` 进行异步HTTP请求下载图片。
#      - 集成PROXY设置，确保图片下载也通过代理进行。
#      - 图片URL处理: 构造原始质量图片URL（添加 `?format=jpg&name=orig` 参数）。
#      - 错误处理: 捕获 `httpx.RequestError` 和 `httpx.HTTPStatusError` 等异常，记录下载失败。
#      - 本地图片路径: 根据发布者名称创建子目录，将图片保存到对应发布者的文件夹中，实现分类存储。
#    - 数据存储: 将抓取到的推文数据（包括图片信息）存储到全局列表 `all_tweet_data` 中。
#    - 唯一发布者存储: 将独特的发布者名称和主页链接存储到全局字典 `unique_authors` 中。
#    - 线程安全: 使用 `asyncio.Lock` 保护全局共享数据结构 (`all_tweet_data` 和 `unique_authors`)，
#      防止多个并发任务同时修改导致数据不一致。
#
# 5. 主执行逻辑 (Main Execution Logic - `main`):
#    - 启动Playwright浏览器。
#    - 调用 `read_urls_from_file` 获取所有目标URL。
#    - 为每个URL创建 `get_illustration` 任务。
#    - 使用 `asyncio.gather(*tasks)` 并发执行所有抓取任务，最大化效率。
#    - 关闭浏览器实例。
#
# 6. Excel数据导出 (Excel Export Logic):
#    - 使用 `openpyxl` 库创建新的Excel工作簿。
#    - 创建两个工作表:
#      - "推文图片信息": 包含所有抓取到的推文的详细数据（任务名称、时间、发布者、内容、图片链接、本地路径等）。
#        - 字段: "任务名称", "发布时间", "发布者", "发布者主页链接", "推文地址", "推文内容", "图片网络地址", "本地图片路径"。
#        - 超链接处理: 为推文地址、发布者主页链接和本地图片路径添加可点击的超链接，方便直接访问。
#        - 本地图片路径超链接: 根据操作系统自动调整路径格式（`file:///` 前缀），并检查文件是否存在。
#      - "唯一发布者信息": 包含所有抓取到的推文中独特的发布者名称和他们的主页链接。
#        - 字段: "发布者名称", "发布者主页链接"。
#        - 超链接处理: 为发布者主页链接添加超链接。
#    - 列宽自适应: 根据列内容的最大长度自动调整列宽，提高可读性。
#    - 目的: 将结构化数据导出，便于后续分析和查阅。
#
# 7. 脚本入口点 (Script Entry Point - `if __name__ == '__main__':`):
#    - 使用 `asyncio.run(main())` 运行主异步函数。
#    - 错误处理:
#      - `KeyboardInterrupt`: 捕获用户中断（Ctrl+C）信号，友好退出。
#      - `Exception`: 捕获所有未处理的异常，记录详细的错误信息和堆栈跟踪，确保程序健壮性。
#    - 清理和自动打开文件 (`finally` 块):
#      - 关闭所有日志处理器，确保日志文件写入完成。
#      - 尝试自动打开生成的日志文件和Excel结果文件，方便用户查看抓取结果和调试信息。
#      - 兼容多操作系统: 根据 `os.name` 或 `os.uname().sysname` 判断操作系统，使用 `os.startfile` (Windows), `open` (macOS), 或 `xdg-open` (Linux) 命令打开文件。
#
# --- 技术栈 ---
# - Python 3.x
# - Playwright (异步浏览器自动化库)
# - BeautifulSoup4 (HTML解析库)
# - httpx (异步HTTP客户端，用于图片下载)
# - openpyxl (Excel文件读写库)
# - asyncio (Python异步编程框架)
# - logging (Python内置日志模块)
# - os, json, datetime, traceback, subprocess (Python标准库)
#
# --- 运行环境要求 ---
# - Python环境已安装。
# - 确保已安装所需的Python库: `pip install playwright beautifulsoup4 httpx openpyxl`
# - 运行 `playwright install` 安装浏览器驱动。
# - 需要一个 `cookies.json` 文件，其中包含X.com的登录cookie，以确保能够访问完整内容。
# - 需要一个 `urlTarget.txt` 文件，其中包含要抓取的X.com个人主页URL，一行一个。
# - 可选配置代理 (`PROXY` 变量)。
#
#
import os
import json
import asyncio
import httpx
from datetime import datetime
import traceback
import logging
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --- FIX FOR NameError: name 'playwright' is not defined ---
# 导入 playwright.async_api 模块并为其创建一个别名，以便在异常处理时可以引用
import playwright.async_api as playwright_api
# 同时从 playwright.async_api 模块直接导入常用的函数和类
from playwright.async_api import async_playwright, expect
# --- END FIX ---

from bs4 import BeautifulSoup


# --- Configuration ---
PROXY = "http://127.0.0.1:10808" # Your proxy address
LOG_DIR = "logs" # Directory for logs
IMAGE_DIR_BASE = "images" # Base directory for images
RESULTS_DIR = "results" # Directory for Excel results
URL_TARGET_FILE = "urlTarget.txt" # New: File to store target URLs

# Create necessary directories if they don't exist
if not os.path.exists(IMAGE_DIR_BASE):
    os.mkdir(IMAGE_DIR_BASE)
if not os.path.exists(LOG_DIR):
    os.mkdir(LOG_DIR)
if not os.path.exists(RESULTS_DIR):
    os.mkdir(RESULTS_DIR)

# Generate a timestamp for unique filenames
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
log_filename = os.path.join(LOG_DIR, f"scraper_log_{timestamp}.txt")
excel_filename = os.path.join(RESULTS_DIR, f"twitter_scrape_results_{timestamp}.xlsx") # Excel filename with timestamp


# --- Logging Setup ---
# Get logger instance
logger = logging.getLogger('twitter_scraper')
logger.setLevel(logging.INFO) # Set minimum logging level to INFO

# Create file handler for logging to a file
file_handler = logging.FileHandler(log_filename, encoding='utf-8')
file_handler.setLevel(logging.INFO)

# Create console handler for logging to console
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# Define log message format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

# Add handlers to the logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)
# --- End Logging Setup ---

# Global list to store data for Excel export
all_tweet_data = []
# Lock for thread-safe access to all_tweet_data when multiple async tasks are running
data_lock = asyncio.Lock()

# Global dictionary to store unique author names and their profile URLs
unique_authors = {}


# --- New helper function to read URLs from file ---
def read_urls_from_file(filepath):
    """Reads URLs from a text file, one URL per line."""
    urls = []
    if not os.path.exists(filepath):
        logger.error(f"Error: URL target file '{filepath}' not found. Please create it with one URL per line.")
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                url = line.strip()
                if url and url.startswith("http"): # Basic validation for URL format
                    urls.append(url)
        if not urls:
            logger.warning(f"Warning: URL target file '{filepath}' is empty or contains no valid URLs.")
        return urls
    except Exception as e:
        logger.error(f"Error reading URLs from '{filepath}': {e}\n{traceback.format_exc()}")
        return []
# --- End new helper function ---


async def get_illustration(context, url):
    # Get the name of the current asynchronous task (e.g., Task-1)
    async_name = asyncio.current_task().get_name()
    # Extract the base username from the URL for easy lookup later (e.g., JIN_HONG_18 from x.com/JIN_HONG_18)
    base_username = url.split('/')[-1].split('?')[0]


    try:
        # Load cookies from cookies.json and add them to the browser context
        with open("cookies.json", "r", encoding="utf-8") as f:
            cookies = json.load(f)

            # Adjust 'sameSite' attribute for Playwright compatibility
            for cookie in cookies:
                cookie_same_site = {'strict': 'Strict', 'Lax': 'lax', 'none': 'None'}.get(cookie.get('sameSite'))
                if cookie_same_site in ['Strict', 'Lax', 'None']:
                    cookie['sameSite'] = cookie_same_site
                else:
                    if 'sameSite' in cookie:
                        del cookie['sameSite'] # Remove if not a valid value for Playwright

            await context.add_cookies(cookies)
        logger.info(f"{async_name} -> Cookies loaded and added to context.")
    except FileNotFoundError:
        logger.error(f"{async_name} -> Error: cookies.json not found. Please create it.")
        return # Exit if cookies file is missing
    except json.JSONDecodeError:
        logger.error(f"{async_name} -> Error: Invalid JSON in cookies.json. Please check the file format.")
        return
    except Exception as e:
        logger.error(f"{async_name} -> Unexpected error loading cookies: {e}")
        return

    # Create a new page in the browser context
    page = await context.new_page()

    try:
        # Navigate to the specified URL, with a timeout
        await page.goto(url, timeout=60000, wait_until="load")
        logger.info(f"{async_name} -> Successfully navigated to {url}")
    except playwright_api.TimeoutError: # --- FIXED: Use playwright_api alias to catch TimeoutError ---
        logger.error(f"{async_name} -> Error: Page.goto timed out for {url} after 60 seconds. Check network or proxy.")
        await page.close()
        return
    except Exception as e:
        logger.error(f"{async_name} -> An unexpected error occurred during navigation to {url}: {e}")
        await page.close()
        return


    # Set to keep track of processed tweet URLs to avoid duplicates
    processed_tweet_urls = set()

    # --- httpx.Proxy setup for image downloads ---
    # Create a httpx.Proxy object if a PROXY is defined.
    # This is preferred for robust proxy handling in httpx to avoid 'dict' object has no attribute 'url' errors.
    proxy_config_for_httpx = None
    if PROXY:
        try:
            proxy_config_for_httpx = httpx.Proxy(PROXY)
            logger.debug(f"httpx proxy configured with: {PROXY}")
        except httpx.UnsupportedProtocol:
            logger.error(f"Unsupported proxy protocol for httpx: {PROXY}")
            proxy_config_for_httpx = None
        except Exception as e:
            logger.error(f"Error creating httpx.Proxy object from {PROXY}: {e}")
            proxy_config_for_httpx = None
    # --- END httpx.Proxy setup ---

    # Loop to scroll and load more content (up to 999 times)
    for x in range(999):
        logger.info(f"{async_name} -> Attempting to scroll and load more content in loop {x}...")
        # Scroll to the bottom of the page
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(2) # Give time for new content to load

        try:
            # Wait for at least one article to be visible, with a timeout
            # --- Added specific catch for AssertionError here ---
            await expect(page.locator('article').nth(0)).to_be_visible(timeout=30000)
        except (playwright_api.TimeoutError, AssertionError) as e: # Catch both TimeoutError and AssertionError
            logger.info(f"{async_name} -> No new articles appeared or first article not visible after scrolling in loop {x}. Breaking loop. Error: {e}")
            break


        # Get all article elements on the page
        articles = page.locator('article')
        article_count = await articles.count()

        if article_count == 0:
            logger.info(f"{async_name} -> No articles found on the page after scrolling. Breaking loop {x}.")
            break

        processed_this_scroll = 0

        # Iterate through each article found on the page
        for i in range(article_count):
            # Initialize a dictionary to store data for the current tweet
            tweet_data = {
                "任务名称": async_name,
                "发布时间": "",
                "发布者": "",
                "发布者主页链接": "",
                "推文地址": "",
                "推文内容": "",
                "图片网络地址": "",
                "本地图片路径": ""
            }
            try:
                current_article = articles.nth(i)
                # Ensure the current article is visible before interacting with it
                # --- Catch AssertionError specifically for individual article visibility ---
                await expect(current_article).to_be_visible(timeout=10000)

                # Parse the article's HTML content with BeautifulSoup
                soup = BeautifulSoup(await current_article.inner_html(timeout=10000), "html.parser")

                # Skip advertisements based on content
                if 'style="text-overflow: unset;">Ad</span>' in str(soup) or soup.find("div", text="Ad") is not None:
                    logger.info(f"{async_name} -> Skipping advertisement on loop {x}, article {i+1}.")
                    continue

                # Find the time element to get publish date and tweet URL
                time_element = soup.find("time")
                if not time_element:
                    logger.warning(f"{async_name} -> Skipping article {i+1} in loop {x}: No time element found (possibly not a standard tweet/ad/unusual content).")
                    continue

                # Extract the tweet URL suffix and construct the full URL
                publish_url_suffix = time_element.find_parent().get("href")
                if not publish_url_suffix:
                    logger.warning(f"{async_name} -> Skipping article {i+1} in loop {x}: No tweet URL suffix found.")
                    continue

                publish_url = "https://x.com" + publish_url_suffix
                tweet_data["推文地址"] = publish_url

                # Skip if tweet URL has already been processed to avoid duplicates
                if publish_url in processed_tweet_urls:
                    continue

                processed_tweet_urls.add(publish_url)
                processed_this_scroll += 1

                logger.info(f"{async_name} -> {'=' * 30}")
                logger.info(f"{async_name} -> Loop {x}, processing tweet {i+1}/{article_count}.")

                # Extract publish time
                publish_time = time_element.get("datetime")
                tweet_data["发布时间"] = datetime.strptime(publish_time, "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%Y年%m月%d日 %H:%M:%S")

                # Extract tweet content
                tweetText = soup.find("div", attrs={"data-testid": "tweetText"})
                publish_content = tweetText.get_text(separator="\n").strip() if tweetText else ""
                tweet_data["推文内容"] = publish_content

                # Extract tweet photos (filter out video thumbnails to download actual images)
                tweetPhoto = soup.find("div", attrs={"data-testid": "tweetPhoto"})
                publish_images = [
                    img.get("src") for img in tweetPhoto.find_all("img")
                    if "video_thumb" not in img.get("src", "") # Exclude video thumbnails
                ] if tweetPhoto else []

                # Extract author name and handle to construct their profile URL
                author_div = soup.find("div", attrs={"data-testid": "User-Name"})
                author_name = ""
                author_handle = ""
                author_profile_url = ""
                if author_div:
                    spans = author_div.find_all('span')
                    if len(spans) >= 1:
                        author_name = spans[0].get_text(strip=True)
                    # Extract author handle and construct profile URL
                    for s in spans:
                        text = s.get_text(strip=True)
                        if text.startswith('@'):
                            author_handle = text
                            author_profile_url = f"https://x.com/{author_handle.lstrip('@')}"
                            break
                    if not author_handle and len(spans) > 1:
                            # Fallback if handle not explicitly marked with '@' but last span looks like it
                            potential_handle = spans[-1].get_text(strip=True)
                            if potential_handle.startswith('@'):
                                author_handle = potential_handle
                                author_profile_url = f"https://x.com/{author_handle.lstrip('@')}"
                            elif url.startswith('https://x.com/') and base_username:
                                # Use the profile from the initial URL if a specific handle isn't found in the tweet content itself
                                author_profile_url = url
                                author_handle = f"@{base_username}" # Infer handle from the initial URL's username


                author = f"{author_name}{' ' + author_handle if author_handle else ''}".strip()
                tweet_data["发布者"] = author
                tweet_data["发布者主页链接"] = author_profile_url

                # Store unique author profile links in the global dictionary for the separate Excel sheet
                if author_name and author_profile_url:
                    async with data_lock: # Protect unique_authors dictionary with an asyncio Lock for thread safety
                        if author_name not in unique_authors:
                            unique_authors[author_name] = author_profile_url


                logger.info(f"{async_name} -> Published: {tweet_data['发布时间']}")
                logger.info(f"{async_name} -> Author: {tweet_data['发布者']}")
                logger.info(f"{async_name} -> Author Profile: {tweet_data['发布者主页链接']}")
                logger.info(f"{async_name} -> Tweet URL: {tweet_data['推文地址']}")
                logger.info(f"{async_name} -> Content: {tweet_data['推文内容']}")
                logger.info(f"{async_name} -> Images: {publish_images}")

                # Process each image for the current tweet
                if not publish_images: # If no images found for this tweet
                    async with data_lock: # Acquire lock before modifying shared list
                        all_tweet_data.append(tweet_data.copy()) # Add a copy of the tweet data (without image details)
                else: # If images are found
                    for image_url in publish_images:
                        image_base_url = image_url[:image_url.rfind("?")]
                        image_name = image_base_url[image_base_url.rfind("/") + 1:]

                        # --- Create author-specific image directory ---
                        # Sanitize author name for folder creation (remove invalid characters)
                        sanitized_author_name = "".join([c for c in author if c.isalnum() or c in (' ', '-', '_')]).strip()
                        if not sanitized_author_name: # Fallback if author name is empty or becomes empty after sanitization
                            sanitized_author_name = "Unknown_Author"

                        # Define the path for the author's image directory
                        author_image_dir = os.path.join(IMAGE_DIR_BASE, sanitized_author_name)
                        if not os.path.exists(author_image_dir):
                            # Create the directory and any necessary parent directories
                            os.makedirs(author_image_dir)

                        # Define the full local path for the image file
                        local_image_path = os.path.join(author_image_dir, f"{image_name}.jpg")
                        # --- End author-specific image directory ---

                        current_image_data = tweet_data.copy() # Make a copy of tweet data for each image
                        current_image_data["图片网络地址"] = image_url # Store original image URL (before ?format=jpg)
                        current_image_data["本地图片路径"] = os.path.abspath(local_image_path) # Store absolute local path


                        try:
                            image_url_orig = image_base_url + "?format=jpg&name=orig" # Construct URL to get original quality image
                            # Download the image using httpx.AsyncClient with proxy
                            async with httpx.AsyncClient(proxy=proxy_config_for_httpx, timeout=30.0) as session:
                                response = await session.get(image_url_orig)
                                response.raise_for_status() # Raise an exception for HTTP errors (e.g., 404, 500)
                                # Save the image content to the local file
                                with open(local_image_path, "wb") as f:
                                    f.write(response.content)
                                logger.info(f"{async_name} -> Downloaded: {local_image_path}")
                        except httpx.RequestError as e:
                            logger.error(f"{async_name} -> Error downloading image {image_url_orig}: {e}")
                        except httpx.HTTPStatusError as e:
                            logger.error(f"{async_name} -> HTTP error downloading image {image_url_orig}: {e.response.status_code} - {e.response.text}")
                        except Exception as e:
                            logger.error(f"{async_name} -> Unexpected error during image download for {image_url_orig}: {e}\n{traceback.format_exc()}")
                        finally:
                            async with data_lock: # Acquire lock before modifying shared list
                                all_tweet_data.append(current_image_data) # Add the image-specific data to the global list

            # --- Catch Playwright TimeoutError and AssertionError for individual articles ---
            except (playwright_api.TimeoutError, AssertionError) as e:
                logger.warning(f"{async_name} -> Playwright Locator Error for article {i+1} in loop {x}: {e}. Skipping this article. Full traceback:\n{traceback.format_exc()}")
            except Exception as e:
                # Catch any other general errors during article processing
                logger.error(f"{async_name} -> General Error processing article {i+1} in loop {x}: {str(e)}\n{traceback.format_exc()}")

        # If no new unique tweets were found after scrolling in this loop (and it's not the first loop), break
        if processed_this_scroll == 0 and x > 0:
            logger.info(f"{async_name} -> No new unique tweets found after scrolling in loop {x}. Exiting loop.")
            break

    await page.close() # Close the Playwright page
    logger.info(f"{async_name} -> Page closed for {url}.")


async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False, # Set to True for headless Browse (runs without a visible browser window)
            proxy={"server": PROXY} if PROXY else None, # Configure proxy for the Playwright browser
            timeout=60000 # Browser launch timeout (in milliseconds)
        )
        context = await browser.new_context() # Create a new browser context (isolated session)

        # --- Optimized section for reading URLs from file ---
        target_urls = read_urls_from_file(URL_TARGET_FILE)
        if not target_urls:
            logger.error(f"No valid URLs found in {URL_TARGET_FILE}. Exiting scraper.")
            await browser.close()
            return # Exit main if no URLs to scrape

        # Create a list of tasks for each URL
        tasks = [get_illustration(context, url) for url in target_urls]

        # Run all scraping tasks concurrently
        await asyncio.gather(*tasks)
        # --- End optimized section ---

        await browser.close() # Close the browser instance after all tasks are done
        logger.info("Browser closed. Script finished scraping data.")

    # --- Excel Export Logic (This block will always run after scraping attempts complete) ---
    wb = Workbook() # Create a new Excel workbook

    # --- Sheet 1: 推文图片信息 (Tweet and Image Info) ---
    ws = wb.active # Get the active worksheet (the first one created by default)
    ws.title = "推文图片信息" # Set the title of the active sheet

    # Define headers for the main tweet/image data sheet
    headers = ["任务名称", "发布时间", "发布者", "发布者主页链接", "推文地址", "推文内容", "图片网络地址", "本地图片路径"]
    ws.append(headers) # Write headers to the first row of the sheet

    # Define font style for hyperlinks (blue, underlined)
    hyperlink_font = Font(color="0000FF", underline="single")

    # Populate the main sheet with collected tweet data
    for row_data in all_tweet_data:
        row = []
        for header in headers:
            # Append data from the dictionary, using an empty string if a key is not found
            row.append(row_data.get(header, ""))
        ws.append(row) # Add the row to the worksheet

        current_row_idx = ws.max_row # Get the current row index for applying hyperlinks

        # Apply hyperlink for "发布者主页链接" (Author Profile Link)
        author_profile_url = row_data.get("发布者主页链接")
        if author_profile_url: # Only apply if URL exists
            cell_author_profile = ws.cell(row=current_row_idx, column=headers.index("发布者主页链接") + 1)
            cell_author_profile.value = author_profile_url # Display the full URL in the cell
            cell_author_profile.hyperlink = author_profile_url # Set the cell's hyperlink
            cell_author_profile.font = hyperlink_font # Apply hyperlink font

        # Apply hyperlink for "推文地址" (Tweet URL)
        tweet_url = row_data.get("推文地址")
        if tweet_url: # Only apply if URL exists
            cell_tweet_url = ws.cell(row=current_row_idx, column=headers.index("推文地址") + 1)
            cell_tweet_url.value = tweet_url
            cell_tweet_url.hyperlink = tweet_url
            cell_tweet_url.font = hyperlink_font

        # Apply hyperlink for "本地图片路径" (Local Image Path)
        local_image_path = row_data.get("本地图片路径")
        if local_image_path and os.path.exists(local_image_path): # Only link if the file actually exists on disk
            # Format local file path for hyperlink (using forward slashes and 'file:///' prefix)
            file_hyperlink_path = local_image_path.replace("\\", "/") # Convert backslashes to forward slashes
            if os.name == 'nt': # For Windows, prepend "file:///" for local file hyperlinks
                file_hyperlink_path = "file:///" + file_hyperlink_path
            else: # For macOS/Linux, prepend "file://"
                file_hyperlink_path = "file://" + file_hyperlink_path

            cell_local_path = ws.cell(row=current_row_idx, column=headers.index("本地图片路径") + 1)
            cell_local_path.value = os.path.basename(local_image_path) # Display just the filename for a cleaner look
            cell_local_path.hyperlink = file_hyperlink_path # Set the local file hyperlink
            cell_local_path.font = hyperlink_font # Apply hyperlink font
        elif local_image_path: # If path exists in data but the file itself doesn't exist on disk
            ws.cell(row=current_row_idx, column=headers.index("本地图片路径") + 1).value = "文件未下载或不存在"


    # Adjust column widths for the main sheet for better readability
    for col_idx, header in enumerate(headers):
        max_length = len(header) # Initialize max length with header's length
        column_letter = get_column_letter(col_idx + 1) # Get Excel column letter (e.g., 'A', 'B', 'C'...)
        # Iterate over all cells in the column to find the maximum content length
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except TypeError:
                pass # Ignore non-string types (e.g., numbers, None)

        adjusted_width = (max_length + 2) * 1.2 # Add padding and a factor for better visual spacing
        if adjusted_width > 100: # Cap maximum column width to prevent excessively wide columns
            adjusted_width = 100
        ws.column_dimensions[column_letter].width = adjusted_width # Set the adjusted column width


    # --- Sheet 2: 唯一发布者信息 (Unique Author Info) ---
    ws_authors = wb.create_sheet("唯一发布者信息") # Create a new sheet specifically for unique authors
    author_headers = ["发布者名称", "发布者主页链接"] # Define headers for the author sheet
    ws_authors.append(author_headers) # Write headers to the authors sheet

    # Populate the unique authors sheet with data from the unique_authors dictionary
    for author_name, profile_url in unique_authors.items():
        row = [author_name, profile_url]
        ws_authors.append(row) # Add author data row

        current_row_idx = ws_authors.max_row # Get current row index for hyperlink
        # Apply hyperlink for "发布者主页链接" in the authors sheet
        cell_author_profile = ws_authors.cell(row=current_row_idx, column=author_headers.index("发布者主页链接") + 1)
        cell_author_profile.value = profile_url # Display the URL
        cell_author_profile.hyperlink = profile_url # Set the hyperlink
        cell_author_profile.font = hyperlink_font # Apply hyperlink font

    # Adjust column widths for the unique authors sheet for readability
    for col_idx, header in enumerate(author_headers):
        max_length = len(header)
        column_letter = get_column_letter(col_idx + 1)
        for cell in ws_authors[column_letter]:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > 100:
            adjusted_width = 100
        ws_authors.column_dimensions[column_letter].width = adjusted_width


    wb.save(excel_filename) # Save the entire workbook to the timestamped Excel file
    logger.info(f"Results saved to Excel: {excel_filename}")
    # --- END Excel Export Logic ---


if __name__ == '__main__':
    try:
        # Run the main asynchronous function
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Script interrupted by user.") # Log if script is interrupted by user (Ctrl+C)
    except Exception as e:
        logger.critical(f"An unhandled error occurred in main: {e}\n{traceback.format_exc()}") # Log any unhandled critical errors
    finally:
        # This 'finally' block ensures cleanup and file opening actions always happen
        # Ensure all log handlers are closed and logs are flushed before exiting
        for handler in logger.handlers[:]: # Iterate over a slice to safely modify list while iterating
            handler.close()
            logger.removeHandler(handler)

        # Automatically open the log file for review
        try:
            if os.path.exists(log_filename):
                print(f"Attempting to open log file: {log_filename}")
                if os.name == 'nt':  # Check if OS is Windows
                    os.startfile(log_filename)
                elif os.uname().sysname == 'Darwin':  # Check if OS is macOS
                    subprocess.run(['open', log_filename])
                else:  # Assume Linux-like system
                    subprocess.run(['xdg-open', log_filename]) # Common command for opening files on Linux
            else:
                print(f"Log file not found: {log_filename}")
        except Exception as e:
            print(f"Error opening log file {log_filename}: {e}")

        # Automatically open the Excel file for review
        try:
            if os.path.exists(excel_filename):
                print(f"Attempting to open Excel file: {excel_filename}")
                if os.name == 'nt':  # Windows
                    os.startfile(excel_filename)
                elif os.uname().sysname == 'Darwin':  # macOS
                    subprocess.run(['open', excel_filename])
                else:  # Linux
                    subprocess.run(['xdg-open', excel_filename])
            else:
                print(f"Excel file not found: {excel_filename}")
        except Exception as e:
            print(f"Error opening Excel file {excel_filename}: {e}")